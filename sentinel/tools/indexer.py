# FILE: tools/indexer.py
import os
import sqlite3
import time
import logging
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
from pathlib import Path
from sentinel.paths import USER_DATA_DIR

DB_FILE = USER_DATA_DIR / "file_index.db"
MAX_FILE_SIZE_MB = 10
THROTTLE_SEC = 0.05

SKIP_DIRS = {
    "node_modules", "venv", ".venv", ".git", "__pycache__",
    "AppData", "site-packages", "Program Files", "Windows", "Library"
}

SKIP_EXT = {'.exe', '.dll', '.bin', '.iso', '.zip', '.tar', '.gz'}


def _get_text_from_file(path, ext):
    """Safely extracts text with memory limits."""
    try:
        size_mb = os.path.getsize(path) / (1024 * 1024)
        if size_mb > MAX_FILE_SIZE_MB:
            return f"[Skipped: Too Large ({size_mb:.1f}MB)]"

        if ext == '.pdf':
            reader = PdfReader(path)
            return " ".join([page.extract_text() for page in reader.pages[:50] if page.extract_text()])

        elif ext == '.docx':
            doc = Document(path)
            return "\n".join([p.text for p in doc.paragraphs])

        elif ext == '.xlsx':
            wb = load_workbook(path, read_only=True, data_only=True)
            content = []
            for sheet in wb.worksheets:
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i > 1000: break
                    row_text = " ".join([str(c) for c in row if c is not None])
                    content.append(row_text)
            return "\n".join(content)

        elif ext == '.csv':
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(10000)
        else:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(20000)

    except Exception:
        return ""


def build_index(root=None, verbose=False):
    """
    Runs smoothly in background without hogging CPU.
    Also queues newly-discovered files into smart_index for semantic embedding.
    """
    if not root:
        root = os.path.expanduser("~")

    # Lazy import to avoid circular dependency at module load time
    try:
        from sentinel.tools.smart_index import index_file as smart_queue
        _smart_available = True
    except Exception:
        _smart_available = False

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('CREATE VIRTUAL TABLE IF NOT EXISTS files USING fts5(path, name, content)')
    c.execute('CREATE TABLE IF NOT EXISTS file_meta (path TEXT PRIMARY KEY, mtime REAL)')

    c.execute("SELECT path, mtime FROM file_meta")
    existing_meta = {row[0]: row[1] for row in c.fetchall()}

    updated = 0

    target_dirs = [
        os.path.join(root, "Documents"),
        os.path.join(root, "Desktop"),
        os.path.join(root, "Downloads")
    ]

    for target in target_dirs:
        if not os.path.exists(target):
            continue

        for r, dirs, files in os.walk(target):
            dirs[:] = [d for d in dirs if d not in SKIP_DIRS and not d.startswith(".")]

            for file in files:
                time.sleep(THROTTLE_SEC)

                ext = os.path.splitext(file)[1].lower()
                if ext in SKIP_EXT:
                    continue
                if ext not in ['.txt', '.md', '.py', '.json', '.pdf', '.docx', '.xlsx', '.csv']:
                    continue

                path = os.path.join(r, file)

                try:
                    stats = os.stat(path)
                    current_mtime = stats.st_mtime

                    if path in existing_meta and existing_meta[path] == current_mtime:
                        continue

                    content = _get_text_from_file(path, ext)
                    if not content or len(content) < 5:
                        continue

                    c.execute("DELETE FROM files WHERE path = ?", (path,))
                    c.execute("INSERT INTO files (path, name, content) VALUES (?,?,?)", (path, file, content))
                    c.execute("INSERT OR REPLACE INTO file_meta (path, mtime) VALUES (?,?)", (path, current_mtime))
                    conn.commit()

                    # Queue into smart_index for background semantic embedding
                    if _smart_available:
                        smart_queue(path)

                    updated += 1

                    if verbose and updated % 10 == 0:
                        print(f"   [Indexer] Indexed: {file} ({updated} total)")

                except Exception:
                    continue

    conn.close()
    if verbose and updated > 0:
        print(f"   [Indexer] Batch complete. {updated} new files added.")
    return f"Indexed {updated} files."


def search_index(query):
    if not os.path.exists(DB_FILE):
        return "Index missing."
    conn = sqlite3.connect(DB_FILE)
    try:
        res = conn.execute(
            "SELECT path, snippet(files, 2, '[', ']', '...', 10) FROM files WHERE files MATCH ? LIMIT 5",
            (query,)
        ).fetchall()
    except Exception:
        return "No matches."
    conn.close()
    if not res:
        return "No text matches found."
    return "\n".join([f"📄 {r[0]}\n   Snippet: \"{r[1]}\"\n" for r in res])